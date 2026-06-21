"""Generate builds.xlsx from extracted RFM build metadata.

Single-shot script. Re-run after touching ROWS below.
Schema is intentionally minimal (6 cols); widen as needed.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = ["build_name", "date", "register", "palette", "pole_position", "score"]

PH = "[PLACEHOLDER]"

ROWS = [
    ("bird_voice_time_chart_v0_1", "2026-06-15",
     "diagram / data-led (架空観測ダイアグラム)",
     "#E8E6DD linen / #1C2A28 ink / #B4593A terracotta",
     "その他 (third lane)", 4),
    ("creative_signal_night", "2026-06-13", PH, PH, PH, PH),
    ("creativevault_hero_firstview_v0_1", "2026-06-13", PH, PH, PH, PH),
    ("creativevault_thought_gallery_rfm_v0_1", "2026-06-14",
     "archive / museum / Letterform (red institutional)",
     "white / ink black / archive red / specimen yellow / deep green",
     "その他 (third lane)", 4),
    ("hospital_year_end_party_2026", "2026-06-13",
     "和モダン seasonal-elegant (sumi+gold)",
     "#211E1B sumi / #ECE5D6 ivory / #C6A260 gold / #BE362C vermilion",
     "その他 (third lane)", 5),
    ("hospital_year_end_party_poster_2026_festive_v0_1", "2026-06-13", PH, PH, PH, PH),
    ("hospital_year_end_party_poster_2026_v0_1", "2026-06-13", PH, PH, PH, PH),
    ("maeda_kekkon_monogatari_v0_1", "2026-06-15",
     "昭和少年劇画 (墨ハッチング+印刷ザラ感)",
     "#0A0A0A 墨 / #F2ECDA 紙地 / #C8311E 朱赤",
     "その他 (gekiga, not loud pop)", 4),
    ("maeda_kekkon_monogatari_v0_2", "2026-06-16",
     "American romance-comic cover (1950s)",
     "newsprint grain + plate offsets",
     "その他 (romance-comic mid-century)", 4),
    ("maeda_kekkon_monogatari_v0_3", "2026-06-16",
     "Golden Age US romance comic (GPT-generated)",
     "romance-comic cover (externally generated)",
     "loud pop (broad comic pole)", 4),
    ("maeda_kekkon_nanoda_rfm_v0_2", "2026-06-17",
     "[PENDING — Gate 0]", PH, PH, PH),
    ("maeda_kekkon_nanoda_v0_1", "2026-06-17",
     "Showa gag-manga (newspaper joke poster) — failed draft",
     "paper noise / halftone / rough borders",
     "loud pop (gag-manga)", 4),
    ("maeda_underwear_wedding_rfm_v0_1", "2026-06-17",
     "[PENDING — output not finalized]", PH, PH, PH),
    ("maeda_wedding_underground_manga_poster_rfm_v0_1", "2026-06-18",
     "near-monochrome underground martial-arts manga wedding poster",
     "#050505 black / #f2f0ea off-white / graphite gray / tiny dried red",
     "その他 (martial-arts manga, image-led)", 4),
    ("memory_garden_key_visual_v0_1", "2026-06-13", PH, PH, PH, PH),
    ("note_header_rfm_journey_v0_1", "2026-06-16",
     "letterpress / wood type specimen (workshop / tool maker)",
     "木目写真 / #fcfaf4 paper / #15110d ink / #b21e16 朱印",
     "その他 (workshop letterpress)", 5),
    ("reference_first_mimic_test_v0_1", "2026-06-12", PH, PH, PH, PH),
    ("reference_first_mimic_test_v0_2", "2026-06-12", PH, PH, PH, PH),
    ("reference_first_mimic_test_v0_3", "2026-06-12", PH, PH, PH, PH),
    ("reference_first_short_film_prototype_v0_1", "2026-06-16", PH, PH, PH, PH),
    ("rfm_codex_test_v0_1", "2026-06-13", PH, PH, PH, PH),
    ("signal_dept_firstview_v0_1", "2026-06-13", PH, PH, PH, PH),
    ("tomorrow_less_boring", "2026-06-13", PH, PH, PH, PH),
    ("ai_users_body_v0_1", "2026-06-21",
     "organic / tactile material (abstract photogram + body x-ray fusion)",
     "#E8E0D2 bone ground / #3A4A5C–#5A6B7D iron-blue flow / #1F2A36 darkest density",
     "その他 (material/photogram, image-led)", 4),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "builds"

    arial = Font(name="Arial", size=11)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1C2A28")
    ph_fill = PatternFill("solid", start_color="FFF4D6")
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    for r in ROWS:
        ws.append(r)

    last_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=last_row,
                            min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.font = arial
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.value == PH or (isinstance(cell.value, str)
                                    and cell.value.startswith("[PENDING")):
                cell.fill = ph_fill

    widths = {"A": 50, "B": 12, "C": 48, "D": 52, "E": 32, "F": 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = "/Users/metcharoba/CreativeVault/70_design_experiments/_inventory/builds.xlsx"
    wb.save(out)
    print(f"wrote {out} — {last_row - 1} rows")


if __name__ == "__main__":
    build()
